from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Exporta DataFrames em um Excel formatado para uso contábil."""

    def export(self, empresa_df, trabalhadores_df, resumo_df, output_path: Path) -> None:
        with self._writer(output_path) as writer:
            empresa_df.to_excel(writer, sheet_name="EMPRESA", index=False)
            trabalhadores_df.to_excel(writer, sheet_name="TRABALHADORES", index=False)
            resumo_df.to_excel(writer, sheet_name="RESUMO", index=False)

        wb = load_workbook(output_path)
        for ws in wb.worksheets:
            self._format_worksheet(ws)
        wb.save(output_path)

    @staticmethod
    def _writer(output_path: Path):
        import pandas as pd

        return pd.ExcelWriter(output_path, engine="openpyxl")

    def _format_worksheet(self, ws) -> None:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

            header = str(column_cells[0].value or "").upper()
            if any(token in header for token in ["TOTAL", "REMUNERACAO", "FGTS", "INSS", "BASE"]):
                for cell in column_cells[1:]:
                    if isinstance(cell.value, (float, int)):
                        cell.number_format = '#,##0.00'
